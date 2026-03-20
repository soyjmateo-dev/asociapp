import openpyxl


def read_excel(file):

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    rows = []

    headers = [cell.value for cell in ws[1]]

    for num_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        data = dict(zip(headers, row))
        data["fila"] = num_fila

        rows.append(data)

    return rows