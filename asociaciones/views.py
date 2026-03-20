from django.http import HttpResponse
from django.contrib.auth import login, logout, get_user_model
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook, load_workbook
from datetime import datetime
from django.contrib import messages
from datetime import date
from core.models import Organizacion
from asociaciones.models import Socio, Actividad, Inscripcion, Pago, Gasto, Patrocinador, Comunicacion, Organismo
from asociaciones.models import TipoComunicacion, ArchivoComunicacion, Contacto, TipoContacto, ArchivoContacto
from asociaciones.models import ItemInventario, CategoriaInventario, ArchivoInventario
from asociaciones.forms import SocioForm, ActividadForm, InscripcionForm, PagoForm, ComunicacionForm, ItemInventarioForm
from asociaciones.forms import ContactoForm, ItemInventarioForm
from django.shortcuts import render, redirect, get_object_or_404
from asociaciones.utils import tenant_login_required
from django.db.models import Q, Sum
from asociaciones.models import Patrocinador
from django import forms
from openpyxl.styles import Font
from django.urls import reverse
from core.excel import formatear_excel

def get_organizacion(slug):
    return get_object_or_404(
        Organizacion,
        slug=slug,
        activa=True
    )

def tenant_login(request, slug):

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)

            return redirect(f"/{slug}/")
    else:
        form = AuthenticationForm()

    return render(request, "tenant_login.html", {
        "form": form,
        "slug": slug
    })


def tenant_home(request, slug):

    organizacion = request.organizacion

    # SOCIOS
    socios = Socio.objects.filter(organizacion=organizacion)

    activos_count = socios.filter(activo=True).count()

    total_deudores = 0
    deuda_total = 0

    for s in socios:
        deuda = s.total_deuda()
        if deuda > 0:
            total_deudores += 1
            deuda_total += deuda

    total_al_corriente = socios.count() - total_deudores


    # FINANZAS

    pagos = Pago.objects.filter(organizacion=organizacion)

    ingresos_cuotas = pagos.filter(
        cuota__isnull=False
    ).aggregate(total=Sum("importe"))["total"] or 0


    ingresos_actividades = pagos.filter(
        inscripcion__isnull=False
    ).aggregate(total=Sum("importe"))["total"] or 0


    ingresos_patrocinios = Patrocinador.objects.filter(
        actividad__organizacion=organizacion
    ).aggregate(total=Sum("aportacion"))["total"] or 0


    gastos = Gasto.objects.filter(
        Q(actividad__organizacion=request.organizacion) |
       Q(actividad__isnull=True)
    ).aggregate(total=Sum("importe"))["total"] or 0


    resultado = ingresos_cuotas + ingresos_actividades + ingresos_patrocinios - gastos


    # ACTIVIDADES

    actividades = Actividad.objects.filter(
        organizacion=organizacion
    ).order_by("fecha")[:5]


    return render(request, "tenant_home.html", {

        "slug": slug,

        "activos_count": activos_count,
        "total_deudores": total_deudores,
        "total_al_corriente": total_al_corriente,
        "deuda_total": deuda_total,

        "ingresos_cuotas": ingresos_cuotas,
        "ingresos_actividades": ingresos_actividades,
        "ingresos_patrocinios": ingresos_patrocinios,
        "gastos": gastos,
        "resultado": resultado,

        "actividades": actividades,
    })

from django.contrib.auth import logout
from django.shortcuts import redirect


def tenant_logout(request, slug):
    logout(request)
    return redirect(f"/{slug}/login/")

from django.urls import reverse

@tenant_login_required
def socios_list(request, slug):
    print("USER:", request.user, request.user.is_authenticated)
    print("SLUG:", slug)

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)
    print("ORG:", organizacion)

    socios = Socio.objects.filter(organizacion=organizacion)
    financiero = request.GET.get("financiero")
    
    # 🔎 BÚSQUEDA
    q = request.GET.get("q", "")
    if q:
        socios = socios.filter(
            Q(nombre__icontains=q) |
            Q(apellidos__icontains=q)
        )

    # 🎛 FILTRO ESTADO
    estado = request.GET.get("estado", "")
    if estado == "activos":
        socios = socios.filter(activo=True)
    elif estado == "inactivos":
        socios = socios.filter(activo=False)

    # 🎛 FILTRO TIPO
    tipo = request.GET.get("tipo", "")

    hoy = date.today()
    fecha_limite = date(hoy.year - 18, hoy.month, hoy.day)

    if tipo == "adultos":
        socios = socios.filter(fecha_nacimiento__lte=fecha_limite)

    elif tipo == "menores":
        socios = socios.filter(fecha_nacimiento__gt=fecha_limite)

    if financiero == "deudores":
        socios = socios.filter(
            cuotas_generadas__pagada=False
        ).distinct()

    elif financiero == "al_corriente":
        socios = socios.exclude(
            cuotas_generadas__pagada=False
        )

    # 🔥 ORDENACIÓN
    orden = request.GET.get("orden", "apellidos")
    direccion = request.GET.get("dir", "asc")

    ordenes_validos = [
        "nombre",
        "apellidos",
        "familia__nombre",
    ]
 
    if orden not in ordenes_validos:
        orden = "apellidos"

    orden_real = orden

    if direccion == "desc":
        orden_real = f"-{orden_real}"

    socios = socios.order_by(orden_real)

    if orden not in ordenes_validos:
        orden = "apellidos"

    else:
        orden_real = orden

    if direccion == "desc":
        orden_real = f"-{orden_real}"

    socios = socios.order_by(orden_real)

    # 📊 CONTADORES GLOBALES
    total = socios.count()
    activos_count = socios.filter(activo=True).count()

    adultos_count = socios.filter(
        fecha_nacimiento__lte=fecha_limite
    ).count()

    menores_count = socios.filter(
        fecha_nacimiento__gt=fecha_limite
    ).count()


    hoy = date.today()
    fecha_limite = date(hoy.year - 18, hoy.month, hoy.day)

    # =========================
    # MÉTRICAS FINANCIERAS
    # =========================

    socios_financieros = Socio.objects.filter(
        organizacion=organizacion
    )

    total_deudores = 0
    deuda_total = 0

    for s in socios_financieros:
        deuda = s.total_deuda()
        if deuda > 0:
            total_deudores += 1
            deuda_total += deuda

    total_al_corriente = socios_financieros.count() - total_deudores
   
    return render(request, "socios_list.html", {
        "socios": socios,
        "slug": slug,
        "q": q,
        "estado": estado,
        "tipo": tipo,
        "total": total,
        "activos_count": activos_count,
        "adultos_count": adultos_count,
        "menores_count": menores_count,
        "financiero": financiero,
        "total_deudores": total_deudores,
        "total_al_corriente": total_al_corriente,
        "deuda_total": deuda_total,
        "orden": orden,
        "direccion": direccion,
        "export_url": reverse("excel_export", args=[slug, "Socio"]),
        "import_url": reverse("socios_importar", args=[slug]),
        "nuevo_url": reverse("socio_create", args=[slug]),
        "deuda_url": reverse("socios_deuda_excel", args=[slug]),
    })

    

def socio_create(request, slug):

    familia_id = request.GET.get("familia")

    if request.method == "POST":
        print("IMPORT EJECUTADO PARA SLUG:", slug)
        organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

        form = SocioForm(
            request.POST or None,
            organizacion=organizacion
        )

        if form.is_valid():
            organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

            socio = form.save(commit=False)
            socio.organizacion = organizacion
            socio.save()

            return redirect(f"/{slug}/socios/")
    else:
        initial_data = {}
        if familia_id:
            initial_data["familia"] = familia_id
        form = SocioForm(initial=initial_data)

    return render(request, "socio_form.html", {
        "form": form,
        "slug": slug
    })


def socio_update(request, slug, pk):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    socio = get_object_or_404(
        Socio,
        pk=pk,
        organizacion=organizacion
    )

    if request.method == "POST":
        form = SocioForm(
            request.POST,
            instance=socio,
            organizacion=organizacion
        )

        if form.is_valid():
            form.save()
            return redirect(f"/{slug}/socios/")
    else:
        form = SocioForm(
            instance=socio,
            organizacion=organizacion
        )

    return render(request, "socio_form.html", {
        "form": form,
        "slug": slug
    })


from django.shortcuts import get_object_or_404, redirect, render

def socio_delete(request, slug, pk):

    organizacion = get_organizacion(slug)

    socio = get_object_or_404(
        Socio,
        pk=pk,
        organizacion=organizacion
    )

    if request.method == "POST":

        confirmacion = request.POST.get("confirmacion", "").strip().lower()
        nombre_real = f"{socio.nombre} {socio.apellidos}".strip().lower()

        if confirmacion == nombre_real:
            socio.delete()
            messages.success(request, "Socio eliminado correctamente.")
            return redirect(f"/{slug}/socios/")
        else:
            messages.error(request, "El nombre introducido no coincide.")

    return render(request, "socio_confirm_delete.html", {
        "socio": socio,
        "slug": slug
    })

from asociaciones.models import Familia
from django import forms


class FamiliaForm(forms.ModelForm):
    class Meta:
        model = Familia
        fields = ["nombre"]

@login_required
def familia_create(request, slug):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    next_url = request.POST.get("next") or request.GET.get("next")

    if request.method == "POST":
        form = FamiliaForm(request.POST)

        if form.is_valid():
            familia = form.save(commit=False)
            familia.organizacion = organizacion
            familia.save()

            if next_url:
                return redirect(f"{next_url}?familia={familia.pk}")

            return redirect(f"/{slug}/socios/")
    else:
        form = FamiliaForm()

    return render(request, "familia_form.html", {
        "form": form,
        "slug": slug,
        "next": next_url,
    })



def familia_update(request, slug, pk):

    organizacion = get_organizacion(slug)

    familia = get_object_or_404(
        Familia,
        pk=pk,
        organizacion=organizacion
    )

    # 👇 ESTA LÍNEA VA AQUÍ
    next_url = request.POST.get("next") or request.GET.get("next")

    if request.method == "POST":
        form = FamiliaForm(request.POST, instance=familia)
        if form.is_valid():
            form.save()
            return redirect(next_url or f"/{slug}/socios/")
    else:
        form = FamiliaForm(instance=familia)

    return render(request, "familia_form.html", {
        "form": form,
        "slug": slug,
        "next": next_url,
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from asociaciones.models import Familia, Socio
from core.models import Organizacion


@login_required
def familia_delete(request, slug, pk):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    familia = get_object_or_404(
        Familia,
        pk=pk,
        organizacion=organizacion
    )

    # 🔒 Verificamos si tiene socios
    if Socio.objects.filter(
        familia=familia,
        organizacion=organizacion
    ).exists():

        messages.error(
            request,
            "No se puede eliminar la familia porque tiene socios asociados."
        )
        return redirect(f"/{slug}/familias/")

    # ✅ Si no tiene socios, se elimina
    familia.delete()

    messages.success(request, "Familia eliminada correctamente.")
    return redirect(f"/{slug}/familias/")

from django.shortcuts import render, get_object_or_404
from core.models import Organizacion
from .models import Familia


def familias_list(request, slug):

    organizacion = get_object_or_404(
        Organizacion,
        slug=slug,
        activa=True
    )

    familias = Familia.objects.filter(
        organizacion=organizacion
    ).order_by("nombre")

    return render(request, "familias_list.html", {
        "familias": familias,
        "slug": slug
    })

def familia_detail(request, slug, pk):

    organizacion = get_organizacion(slug)

    familia = get_object_or_404(
        Familia,
        pk=pk,
        organizacion=organizacion
    )
    socios = familia.socio_set.all()

    return render(request, "familia_detail.html", {
        "familia": familia,
        "socios": socios,
        "slug": slug
    })

from django.apps import apps

from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from core.models import Organizacion
from asociaciones.models import Socio

@login_required
def export_socios_excel(request, slug):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    socios = Socio.objects.filter(
        organizacion=organizacion
    ).order_by("apellidos")

    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"

    fields = [
        f for f in Socio._meta.fields
        if f.name not in ["id", "organizacion"]
    ]

    headers = [f.name for f in fields]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws.freeze_panes = "A2"

    for s in socios:
        row = []
        for f in fields:
            value = getattr(s, f.name)

            if value is None:
                row.append("")
            elif hasattr(value, "strftime"):
                row.append(value.strftime("%Y-%m-%d"))
            elif f.many_to_one:
                row.append(str(value))
            else:
                row.append(value)

        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=socios_{slug}.xlsx"
    formatear_excel(ws)
    wb.save(response)
    return response

@login_required
def descargar_plantilla_socios(request, slug):

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Socios"

    fields = [
        f for f in Socio._meta.fields
        if f.name not in ["id", "organizacion"]
    ]

    headers = [f.name for f in fields]
    ws.append(headers)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=plantilla_socios_{slug}.xlsx"
    formatear_excel(ws)
    wb.save(response)
    return response

from django.db import transaction

@tenant_login_required
def import_socios_excel(request, slug):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    if request.method == "POST" and request.FILES.get("archivo"):

        archivo = request.FILES["archivo"]
        wb = load_workbook(archivo)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        errores = []
        filas_validadas = []

        # -----------------------------
        # VALIDAR TODO
        # -----------------------------

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            data = dict(zip(headers, row))

            try:

                fecha = data.get("fecha_nacimiento")

                if isinstance(fecha, datetime):
                    fecha = fecha.date()

                elif isinstance(fecha, str) and fecha:
                    try:
                        fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                    except ValueError:
                        fecha = datetime.strptime(fecha, "%d/%m/%Y").date()

                else:
                    fecha = None

                familia_obj = None
                nombre_familia = data.get("familia")

                # solo guardar datos para después
                filas_validadas.append({
                    "nombre": data.get("nombre"),
                    "apellidos": data.get("apellidos"),
                    "fecha_nacimiento": fecha,
                    "familia": nombre_familia,
                    "email": data.get("email"),
                    "telefono": data.get("telefono"),
                    "row": row_num
                })

            except Exception as e:

                errores.append({
                    "fila": row_num,
                    "error": str(e)
                })

        # -----------------------------
        # SI HAY ERRORES → CANCELAR
        # -----------------------------

        if errores:

            return render(request, "socios_import.html", {
                "slug": slug,
                "errores": errores
            })

        # -----------------------------
        # IMPORTAR TODO EN TRANSACCIÓN
        # -----------------------------

        creados = 0

        with transaction.atomic():

            for fila in filas_validadas:

                familia_obj = None

                if fila["familia"]:

                    familia_obj, _ = Familia.objects.get_or_create(
                        nombre=fila["familia"],
                        organizacion=organizacion
                    )

                Socio.objects.create(
                    organizacion=organizacion,
                    nombre=fila["nombre"],
                    apellidos=fila["apellidos"],
                    fecha_nacimiento=fila["fecha_nacimiento"],
                    familia=familia_obj,
                    telefono=fila["telefono"],
                    email=fila["email"],
                )

                creados += 1

        messages.success(request, f"{creados} socios importados correctamente")

        return redirect(f"/{slug}/socios/")

    return render(request, "socios_import.html", {"slug": slug})

from django.db import transaction
from datetime import date
from asociaciones.models import Cuota, ConfiguracionCuota

@login_required
def generar_cuotas_anuales(request, slug):

    año = date.today().year

    organizacion = get_organizacion(slug)

    socios = Socio.objects.filter(
        organizacion=organizacion,
        activo=True
    )

    creadas = 0
    existentes = 0

    with transaction.atomic():

        for socio in socios:

            # Determinar tipo
            tipo = "menor" if socio.es_menor else "adulto"

            config = ConfiguracionCuota.objects.filter(
                tipo=tipo,
                activa=True
            ).first()

            if not config:
                continue

            cuota, created = Cuota.objects.get_or_create(
                socio=socio,
                año=año,
                defaults={
                    "importe": config.importe
                }
            )

            if created:
                creadas += 1
            else:
                existentes += 1

    messages.success(
        request,
        f"Cuotas generadas: {creadas} | Ya existentes: {existentes}"
    )

    return redirect(f"/{slug}/socios/")

from django.http import JsonResponse
import json

@login_required
def familia_ajax_crear(request, slug):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    data = json.loads(request.body)
    nombre = data.get("nombre")

    familia = Familia.objects.create(
        nombre=nombre,
        organizacion=organizacion
    )

    return JsonResponse({
        "id": familia.id,
        "nombre": familia.nombre
    })

from django.shortcuts import render, get_object_or_404
from core.models import Organizacion
from .models import Cuota


def cuotas_list(request, slug):

    organizacion = get_object_or_404(
        Organizacion,
        slug=slug,
        activa=True
    )

    cuotas = Cuota.objects.filter(
        organizacion=organizacion
    ).order_by("-año")

    return render(request, "core/cuotas_list.html", {
        "cuotas": cuotas,
        "slug": slug,
    })

from django import forms
from datetime import date
from .models import Cuota
from core.models import Organizacion


class CuotaForm(forms.ModelForm):

    def __init__(self, *args, organizacion=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.organizacion = organizacion

    class Meta:
        model = Cuota
        fields = [
            "nombre",
            "año",
            "tipo",
            "importe_adulto",
            "importe_menor",
            "fecha_vencimiento",
        ]

    def clean(self):
        cleaned_data = super().clean()
        tipo = cleaned_data.get("tipo")
        año = cleaned_data.get("año")

        if tipo == "anual" and self.organizacion and año:
            existe = Cuota.objects.filter(
                organizacion=self.organizacion,
                tipo="anual",
                año=año
            ).exists()

            if existe:
                raise forms.ValidationError(
                    f"Ya existe una cuota anual para el año {año}."
                )

        return cleaned_data

def cuota_create(request, slug):

    organizacion = get_object_or_404(Organizacion, slug=slug, activa=True)

    if request.method == "POST":
        form = CuotaForm(request.POST, organizacion=organizacion)
        if form.is_valid():
            cuota = form.save(commit=False)
            cuota.estado = "borrador"
            cuota.organizacion = organizacion  # 👈 ESTO ES CLAVE SI TU MODELO LO USA
            cuota.save()
            return redirect("cuotas_list", slug=slug)
    else:
        form = CuotaForm(
            initial={"año": date.today().year},
            organizacion=organizacion
        )

    return render(request, "core/cuota_form.html", {
        "form": form,
        "slug": slug,
    })

from django.shortcuts import get_object_or_404, redirect
from .models import Cuota, generar_cuotas_para_definicion


def cuota_activar(request, slug, cuota_id):

    organizacion = get_organizacion(slug)

    cuota = get_object_or_404(
        Cuota,
        id=cuota_id,
        organizacion=organizacion
    )

    if cuota.estado != "borrador":
        messages.error(request, "Solo se pueden activar cuotas en borrador.")
        return redirect("cuotas_list", slug=slug)

    cuota.estado = "activa"
    cuota.save()

    generar_cuotas_para_definicion(cuota)

    messages.success(request, "Cuota activada y cuotas generadas correctamente.")

    return redirect("cuotas_list", slug=slug)

from .models import Cuota, CuotaSocio
from django.shortcuts import get_object_or_404


def cuota_socios_list(request, slug, cuota_id):

    organizacion = get_object_or_404(
        Organizacion,
        slug=slug,
        activa=True
    )

    cuota = get_object_or_404(
        Cuota,
        id=cuota_id,
        organizacion=organizacion
    )

    cuotas_socio = CuotaSocio.objects.filter(
        cuota=cuota,
        cuota__organizacion=organizacion,
        socio__organizacion=organizacion
    ).select_related("socio")

    return render(request, "core/cuota_socios_list.html", {
        "cuota": cuota,
        "cuotas_socio": cuotas_socio,
        "slug": slug,
    })

from django.contrib.auth.views import PasswordChangeView
from asociaciones.utils import tenant_login_required


class TenantPasswordChangeView(PasswordChangeView):
    template_name = "registration/password_change_form.html"

    def get_success_url(self):
        return f"/{self.kwargs['slug']}/password-change/done/"
    

from .models import Actividad
from asociaciones.utils import tenant_login_required


@tenant_login_required
def actividades_list(request, slug):

    actividades = Actividad.objects.filter(
        organizacion=request.organizacion
    )

    return render(
        request,
        "actividades_list.html",
        {
            "actividades": actividades,
            "slug": slug,
        }
    )

from .forms import ActividadForm


@tenant_login_required
def actividad_create(request, slug):

    if request.method == "POST":
        form = ActividadForm(request.POST)

        if form.is_valid():
            actividad = form.save(commit=False)

            # 🔐 Asignación automática multi-tenant
            actividad.organizacion = request.organizacion

            actividad.save()

            return redirect(f"/{slug}/actividades/")

    else:
        form = ActividadForm()

    return render(
        request,
        "core/actividad_form.html",  # ← ESTA ES LA CLAVE
        {
            "form": form,
            "slug": slug,
        }
    )

from django.shortcuts import get_object_or_404, redirect


def actividad_update(request, slug, pk):
    actividad = get_object_or_404(
        Actividad,
        pk=pk,
        organizacion=request.organizacion
    )

    if request.method == "POST":
        form = ActividadForm(request.POST, instance=actividad)
        if form.is_valid():
            form.save()
            messages.success(request, "Actividad actualizada correctamente")
            return redirect("actividades_list", slug=slug)
    else:
        form = ActividadForm(instance=actividad)

    return render(request, "core/actividad_form.html", {
        "form": form,
        "slug": slug
    })


def actividad_delete(request, slug, pk):
    actividad = get_object_or_404(
        Actividad,
        pk=pk,
        organizacion=request.organizacion
    )

    actividad.delete()
    messages.success(request, "Actividad eliminada")
    return redirect("actividades_list", slug=slug)

@tenant_login_required
def inscripcion_create(request, slug, actividad_id):

    actividad = get_object_or_404(
        Actividad,
        id=actividad_id,
        organizacion=request.organizacion
    )

    inscritos = actividad.inscripciones.select_related("socio")

    if request.method == "POST":
        form = InscripcionForm(
            request.POST,
            organizacion=request.organizacion,
            actividad=actividad
        )

        if form.is_valid():

            inscripcion = form.save(commit=False)
            inscripcion.actividad = actividad
            inscripcion.save()

            messages.success(request, "Inscripción realizada correctamente")

            return redirect(
                "inscripcion_create",
                slug=slug,
                actividad_id=actividad.id
            )

    else:
        form = InscripcionForm(
            organizacion=request.organizacion,
            actividad=actividad
        )

    return render(
        request,
        "core/inscripcion_form.html",
        {
            "form": form,
            "actividad": actividad,
            "inscritos": inscritos,
            "slug": slug,
        }
    )

@tenant_login_required
def inscripcion_delete(request, slug, inscripcion_id):

    inscripcion = get_object_or_404(
        Inscripcion,
        id=inscripcion_id,
        actividad__organizacion=request.organizacion
    )

    actividad_id = inscripcion.actividad.id
    inscripcion.delete()

    messages.success(request, "Inscripción eliminada")

    return redirect("actividad_inscritos", slug=slug, actividad_id=actividad_id)

@tenant_login_required
def actividad_inscritos(request, slug, actividad_id):

    actividad = get_object_or_404(
        Actividad,
        id=actividad_id,
        organizacion=request.organizacion
    )

    inscripciones = actividad.inscripciones.select_related("socio")

    total_recaudado = actividad.ingresos_totales()

    return render(
        request,
        "core/actividad_inscritos.html",
        {
            "actividad": actividad,
            "inscripciones": inscripciones,
            "total_recaudado": total_recaudado,
            "slug": slug,
        }
    )

@tenant_login_required
def pago_familiar(request, slug, familia_id, cuota_id):

    familia = get_object_or_404(
        Familia,
        id=familia_id,
        organizacion=request.organizacion
    )

    cuota = get_object_or_404(
        Cuota,
        id=cuota_id,
        organizacion=request.organizacion
    )

    socios = familia.socio_set.filter(activo=True)

    for socio in socios:
        cuota_socio = socio.cuotas_generadas.filter(
            cuota=cuota,
            pagada=False
        ).first()

        if cuota_socio:
            Pago.objects.create(
                socio=socio,
                familia=familia,
                cuota=cuota,
                organizacion=request.organizacion,
                importe=cuota_socio.importe,
                metodo="efectivo"
            )

    messages.success(request, "Pago familiar registrado correctamente")

    return redirect("cuota_socios_list", slug=slug, cuota_id=cuota.id)

@tenant_login_required
def pagos_list(request, slug):

    pagos = Pago.objects.filter(
        organizacion=request.organizacion
    ).select_related(
        "socio", "cuota", "inscripcion", "familia"
    ).order_by("-fecha", "-id")

    return render(
        request,
        "core/pagos_list.html",
        {
            "pagos": pagos,
            "slug": slug,
        }
    )

@tenant_login_required
def pago_create(request, slug):

    print("===== PAGO_CREATE EJECUTADO =====")
    print("METHOD:", request.method)

    if request.method == "POST":
        print("POST DATA:", request.POST)

        form = PagoForm(request.POST, organizacion=request.organizacion)

        print("FORM VALID?:", form.is_valid())
        print("FORM ERRORS:", form.errors)

        if form.is_valid():

            pago = form.save(commit=False)
            pago.organizacion = request.organizacion

            actividad = form.cleaned_data.get("actividad")
            socio = form.cleaned_data.get("socio")

            if actividad:
                inscripcion = Inscripcion.objects.get(
                    actividad=actividad,
                    socio=socio
                )
                pago.inscripcion = inscripcion

            pago.save()

            return redirect("pagos_list", slug=slug)
    else:
        socio_id = request.GET.get("socio")
        initial = {}

        if socio_id:
            initial["socio"] = socio_id

        form = PagoForm(
            organizacion=request.organizacion,
            initial=initial
        )

    return render(
        request,
        "core/pago_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

from django.http import JsonResponse
from asociaciones.models import Cuota

@tenant_login_required
def cuota_importe(request, slug, cuota_id):
    cuota = get_object_or_404(
        Cuota,
        id=cuota_id,
        organizacion=request.organizacion
    )

    return JsonResponse({
        "importe": str(cuota.importe)
    })

from django.http import JsonResponse
from django.db.models import Sum
from asociaciones.models import CuotaSocio, Inscripcion

@tenant_login_required
def obtener_importe(request, slug):

    cuota_id = request.GET.get("cuota")
    socio_id = request.GET.get("socio")
    familia_id = request.GET.get("familia")
    actividad_id = request.GET.get("actividad")

    # 🔹 CUOTA INDIVIDUAL
    if cuota_id and socio_id:
        cuota = Cuota.objects.get(id=cuota_id)
        socio = Socio.objects.get(id=socio_id)

        cuota_socio = socio.cuotas_generadas.filter(cuota=cuota).first()

        if cuota_socio:
            return JsonResponse({"importe": str(cuota_socio.importe)})

    # 🔹 CUOTA FAMILIAR
    if cuota_id and familia_id:
        familia = Familia.objects.get(id=familia_id)

        total = 0
        for socio in familia.socio_set.all():
            cuota_socio = socio.cuotas_generadas.filter(
                cuota_id=cuota_id,
                pagada=False
            ).first()

            if cuota_socio:
                total += cuota_socio.importe

        return JsonResponse({"importe": str(total)})

    # 🔹 ACTIVIDAD
    if actividad_id and socio_id:
        actividad = Actividad.objects.get(
            id=actividad_id,
            organizacion=request.organizacion
        )

        socio = Socio.objects.get(id=socio_id)

        if socio.es_menor:
            importe = actividad.coste_menor
        else:
            importe = actividad.coste_adulto

        return JsonResponse({"importe": str(importe)})

    return JsonResponse({"importe": "0"})

@tenant_login_required
def obtener_socios_actividad(request, slug):

    actividad_id = request.GET.get("actividad")

    if not actividad_id:
        return JsonResponse({"socios": []})

    # Socios inscritos
    inscripciones = Inscripcion.objects.filter(
        actividad_id=actividad_id,
        actividad__organizacion=request.organizacion
    ).select_related("socio")

    socios_ids_inscritos = [i.socio.id for i in inscripciones]

    # Socios que ya pagaron esa actividad
    socios_ids_pagados = Pago.objects.filter(
        actividad_id=actividad_id,
        socio__organizacion=request.organizacion
    ).values_list("socio_id", flat=True)

    # Excluir los que ya pagaron
    socios_pendientes = inscripciones.exclude(
        socio_id__in=socios_ids_pagados
    )

    socios = [
        {
            "id": i.socio.id,
            "nombre": str(i.socio)
        }
        for i in socios_pendientes
    ]

    return JsonResponse({"socios": socios})

@tenant_login_required
def obtener_socios_cuota(request, slug):

    cuota_id = request.GET.get("cuota")

    if not cuota_id:
        return JsonResponse({"socios": [], "familias": []})

    organizacion = request.organizacion

    cuota = Cuota.objects.filter(
        id=cuota_id,
        organizacion=organizacion,
        estado="activa"
    ).first()

    if not cuota:
        return JsonResponse({"socios": [], "familias": []})

    # 🔵 SOCIOS QUE DEBEN ESA CUOTA
    cuotas_pendientes = CuotaSocio.objects.filter(
        cuota=cuota,
        pagada=False,
        socio__organizacion=organizacion
    ).select_related("socio")

    socios_data = [
        {
            "id": cs.socio.id,
            "nombre": str(cs.socio)
        }
        for cs in cuotas_pendientes
    ]

    # 🟢 FAMILIAS QUE TIENEN ALGÚN SOCIO CON CUOTA PENDIENTE
    familias_ids = cuotas_pendientes.values_list(
        "socio__familia_id",
        flat=True
    ).distinct()

    familias = Familia.objects.filter(
        id__in=familias_ids,
        organizacion=organizacion
    )

    familias_data = [
        {
            "id": f.id,
            "nombre": f.nombre
        }
        for f in familias
    ]

    return JsonResponse({
        "socios": socios_data,
        "familias": familias_data
    })

from django import forms

class GastoForm(forms.ModelForm):
    class Meta:
        model = Gasto
        fields = [
            "fecha",
            "concepto",
            "importe",
            "metodo_pago",
            "actividad",
            "archivo",
            "observaciones",
        ]

        widgets = {
            "fecha": forms.DateInput(
                attrs={"type": "date", "class": "form-control"}
            ),
            "concepto": forms.TextInput(
                attrs={"class": "form-control"}
            ),
            "importe": forms.NumberInput(
                attrs={"class": "form-control", "step": "0.01"}
            ),
            "metodo_pago": forms.Select(
                attrs={"class": "form-select"}
            ),
            "actividad": forms.Select(
                attrs={"class": "form-select"}
            ),
            "archivo": forms.ClearableFileInput(
                attrs={
                    "class": "form-control",
                    "accept": "image/*,application/pdf"
                }
            ),
            "observaciones": forms.Textarea(
                attrs={
                    "class": "form-control",
                    "rows": 2,
                    "style": "resize:none;"
                }
            ),
        }
@tenant_login_required

def gastos_list(request, slug):

    gastos = Gasto.objects.filter(
        Q(actividad__organizacion=request.organizacion) |
        Q(actividad__isnull=True)
    ).order_by("-fecha")

    return render(
        request,
        "core/gastos_list.html",
        {
            "gastos": gastos,
            "slug": slug
        }
    )

def gasto_delete(request, slug, pk):

    gasto = get_object_or_404(Gasto, pk=pk)

    if request.method == "POST":
        gasto.delete()

    return redirect("gastos_list", slug=slug)

@tenant_login_required
def gasto_create(request, slug):

    if request.method == "POST":
        form = GastoForm(request.POST, request.FILES)
        if form.is_valid():
            gasto = form.save()
            return redirect("gastos_list", slug=slug)
    else:
        form = GastoForm()

    return render(
        request,
        "core/gasto_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

class PatrocinadorForm(forms.ModelForm):
    class Meta:
        model = Patrocinador
        fields = [
            "empresa",
            "nombre_contacto",
            "telefono",
            "email",
            "actividad",
            "aportacion",
            "año",
            "logotipo",
            "documento",
            "activo",
            "observaciones",
        ]

        widgets = {
            "empresa": forms.TextInput(attrs={"class": "form-control"}),
            "nombre_contacto": forms.TextInput(attrs={"class": "form-control"}),
            "telefono": forms.TextInput(attrs={"class": "form-control"}),
            "email": forms.EmailInput(attrs={"class": "form-control"}),
            "actividad": forms.Select(attrs={"class": "form-select"}),
            "aportacion": forms.NumberInput(attrs={"class": "form-control", "step": "0.01"}),
            "año": forms.NumberInput(attrs={"class": "form-control"}),
            "logotipo": forms.ClearableFileInput(attrs={"class": "form-control", "accept": "image/*"}),
            "documento": forms.ClearableFileInput(attrs={"class": "form-control",
            "accept": ".pdf,.doc,.docx,.xls,.xlsx,image/*"}),
            "activo": forms.CheckboxInput(attrs={"class": "form-check-input"}),
            "observaciones": forms.Textarea(attrs={"class": "form-control", "rows": 2}),
        }

@tenant_login_required
def patrocinadores_list(request, slug):

    patrocinadores = Patrocinador.objects.filter(
        actividad__organizacion=request.organizacion
    ).select_related("actividad").order_by("-año", "empresa")

    return render(
        request,
        "core/patrocinadores_list.html",
        {
            "patrocinadores": patrocinadores,
            "slug": slug,
        }
    )

@tenant_login_required
def patrocinador_create(request, slug):

    if request.method == "POST":
        form = PatrocinadorForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("patrocinadores_list", slug=slug)
    else:
        form = PatrocinadorForm()

    return render(
        request,
        "core/patrocinador_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

from django.shortcuts import get_object_or_404
from django.db.models import Sum
from asociaciones.models import Actividad, Pago, Gasto, Patrocinador
from .models import Inscripcion

@tenant_login_required
def actividad_detail(request, slug, pk):

    actividad = get_object_or_404(
        Actividad,
        pk=pk,
        organizacion=request.organizacion
    )

    inscripciones = Inscripcion.objects.filter(
        actividad=actividad
    ).select_related("socio")

    context = {
        "actividad": actividad,
        "inscripciones": inscripciones,
        "ingresos": actividad.ingresos_totales(),
        "patrocinios": actividad.patrocinios.aggregate(total=Sum("aportacion"))["total"] or 0,
        "gastos": actividad.total_gastos(),
        "resultado": actividad.beneficio(),
        "slug": slug
    }

    return render(request, "core/actividad_detail.html", context)

@tenant_login_required
def patrocinador_detail(request, slug, pk):

    patrocinador = get_object_or_404(
        Patrocinador,
        pk=pk,
        actividad__organizacion=request.organizacion
    )

    return render(
        request,
        "core/patrocinador_detail.html",
        {
            "patrocinador": patrocinador,
            "slug": slug,
        }
    )

@tenant_login_required
def patrocinador_update(request, slug, pk):

    patrocinador = get_object_or_404(
        Patrocinador,
        pk=pk,
        actividad__organizacion=request.organizacion
    )

    if request.method == "POST":
        form = PatrocinadorForm(request.POST, request.FILES, instance=patrocinador)
        if form.is_valid():
            form.save()
            return redirect("patrocinador_detail", slug=slug, pk=pk)
    else:
        form = PatrocinadorForm(instance=patrocinador)

    return render(
        request,
        "core/patrocinador_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

@tenant_login_required
def comunicaciones_list(request, slug):

    organizacion = request.organizacion

    comunicaciones = Comunicacion.objects.filter(
        organizacion=organizacion
    ).select_related("organismo", "tipo")

    # 🔎 BUSCADOR
    q = request.GET.get("q", "")
    if q:
        comunicaciones = comunicaciones.filter(
            Q(asunto__icontains=q) |
            Q(descripcion__icontains=q)
        )

    # 🏛 FILTRO ORGANISMO
    organismo_id = request.GET.get("organismo")
    if organismo_id:
        comunicaciones = comunicaciones.filter(organismo_id=organismo_id)

    # 🏷 FILTRO TIPO
    tipo_id = request.GET.get("tipo")
    if tipo_id:
        comunicaciones = comunicaciones.filter(tipo_id=tipo_id)

    # 📅 FILTRO AÑO
    año = request.GET.get("año")
    if año:
        comunicaciones = comunicaciones.filter(fecha__year=año)

    comunicaciones = comunicaciones.order_by("-fecha")

    # 📂 AGRUPAR POR AÑO
    comunicaciones_por_año = {}
    for c in comunicaciones:
        año = c.fecha.year
        comunicaciones_por_año.setdefault(año, []).append(c)

    # Datos para filtros
    organismos = Organismo.objects.filter(organizacion=organizacion)
    tipos = TipoComunicacion.objects.filter(organizacion=organizacion)

    años_disponibles = (
        Comunicacion.objects
        .filter(organizacion=organizacion)
        .dates("fecha", "year")
    )

    return render(
        request,
        "core/comunicaciones_list.html",
        {
            "slug": slug,
            "comunicaciones_por_año": comunicaciones_por_año,
            "organismos": organismos,
            "tipos": tipos,
            "años_disponibles": años_disponibles,
            "q": q,
            "organismo_id": organismo_id,
            "tipo_id": tipo_id,
            "año_filtro": año,
        }
    )

@tenant_login_required
def comunicacion_create(request, slug):

    if request.method == "POST":
        form = ComunicacionForm(
            request.POST,
            organizacion=request.organizacion
        )

        if form.is_valid():
            comunicacion = form.save(commit=False)
            comunicacion.organizacion = request.organizacion
            comunicacion.creada_por = request.user
            comunicacion.save()

            # 🔥 GUARDAR ARCHIVOS MULTIPLES
            for archivo in request.FILES.getlist("archivos"):
                ArchivoComunicacion.objects.create(
                    comunicacion=comunicacion,
                    archivo=archivo
                )

            return redirect("comunicaciones_list", slug=slug)

    else:
        form = ComunicacionForm(
            organizacion=request.organizacion
        )

    return render(
        request,
        "core/comunicacion_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

@tenant_login_required
def organismos_list(request, slug):

    organismos = Organismo.objects.filter(
        organizacion=request.organizacion
    )

    return render(
        request,
        "core/organismos_list.html",
        {
            "organismos": organismos,
            "slug": slug,
        }
    )

@tenant_login_required
def organismo_create(request, slug):

    if request.method == "POST":
        form = OrganismoForm(request.POST)

        if form.is_valid():
            organismo = form.save(commit=False)
            organismo.organizacion = request.organizacion
            organismo.save()

            return redirect("organismos_list", slug=slug)
    else:
        form = OrganismoForm()

    return render(
        request,
        "core/organismo_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

@tenant_login_required
def tipo_comunicacion_create(request, slug):

    if request.method == "POST":
        form = TipoComunicacionForm(request.POST)

        if form.is_valid():
            tipo = form.save(commit=False)
            tipo.organizacion = request.organizacion
            tipo.save()

            return redirect("tipos_comunicacion_list", slug=slug)
    else:
        form = TipoComunicacionForm()

    return render(
        request,
        "core/tipo_comunicacion_form.html",
        {
            "form": form,
            "slug": slug,
        }
    )

from django.http import JsonResponse
from django.views.decorators.http import require_POST

@tenant_login_required
@require_POST
def organismo_ajax_create(request, slug):

    nombre = request.POST.get("nombre")

    if not nombre:
        return JsonResponse({"error": "Nombre requerido"}, status=400)

    organismo = Organismo.objects.create(
        nombre=nombre,
        organizacion=request.organizacion,
        activo=True,
    )

    return JsonResponse({
        "id": organismo.id,
        "nombre": organismo.nombre
    })

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from asociaciones.models import TipoComunicacion

@tenant_login_required
@require_POST
def tipo_comunicacion_ajax_create(request, slug):

    nombre = request.POST.get("nombre")

    if not nombre:
        return JsonResponse({"error": "Nombre requerido"}, status=400)

    tipo = TipoComunicacion.objects.create(
        nombre=nombre,
        organizacion=request.organizacion
    )

    return JsonResponse({
        "id": tipo.id,
        "nombre": tipo.nombre
    })

@tenant_login_required
def comunicacion_detail(request, slug, pk):

    comunicacion = get_object_or_404(
        Comunicacion,
        pk=pk,
        organizacion=request.organizacion
    )

    archivos = comunicacion.archivos.all()

    return render(
        request,
        "core/comunicacion_detail.html",
        {
            "comunicacion": comunicacion,
            "archivos": archivos,
            "slug": slug,
        }
    )

@tenant_login_required
def comunicacion_update(request, slug, pk):

    comunicacion = get_object_or_404(
        Comunicacion,
        pk=pk,
        organizacion=request.organizacion
    )

    if request.method == "POST":
        form = ComunicacionForm(
            request.POST,
            instance=comunicacion,
            organizacion=request.organizacion
        )

        if form.is_valid():
            comunicacion = form.save()

            # 👇 NUEVOS ARCHIVOS
            for archivo in request.FILES.getlist("archivos"):
                ArchivoComunicacion.objects.create(
                    comunicacion=comunicacion,
                    archivo=archivo
                )

            messages.success(request, "Comunicación actualizada correctamente")
            return redirect("comunicaciones_list", slug=slug)

    else:
        form = ComunicacionForm(
            instance=comunicacion,
            organizacion=request.organizacion
        )

    return render(
        request,
        "core/comunicacion_form.html",
        {
            "form": form,
            "slug": slug,
            "comunicacion": comunicacion
        }
    )

from django.views.decorators.http import require_POST

@tenant_login_required
@require_POST
def archivo_comunicacion_delete(request, slug, pk):

    archivo = get_object_or_404(
        ArchivoComunicacion,
        pk=pk,
        comunicacion__organizacion=request.organizacion
    )

    comunicacion_id = archivo.comunicacion.id

    # 🔥 Borra el archivo físico
    archivo.archivo.delete(save=False)

    # 🔥 Borra el registro
    archivo.delete()

    messages.success(request, "Archivo eliminado correctamente.")

    return redirect("comunicacion_update", slug=slug, pk=comunicacion_id)

@tenant_login_required
def contactos_list(request, slug):

    organizacion = request.organizacion

    q = request.GET.get("q")
    tipo = request.GET.get("tipo")

    contactos = Contacto.objects.filter(
        organizacion=organizacion
    ).select_related("tipo").order_by("nombre")

    if q:
        contactos = contactos.filter(
            nombre__icontains=q
        ) | contactos.filter(
            apellidos__icontains=q
        ) | contactos.filter(
            empresa__icontains=q
        )

    if tipo:
        contactos = contactos.filter(tipo_id=tipo)

    tipos = TipoContacto.objects.filter(organizacion=organizacion)

    return render(
        request,
        "core/contactos_list.html",
        {
            "contactos": contactos,
            "tipos": tipos,
            "slug": slug,
        }
    )

@tenant_login_required
def contacto_create(request, slug):

    organizacion = request.organizacion

    if request.method == "POST":

        form = ContactoForm(request.POST, request.FILES)

        if form.is_valid():

            contacto = form.save(commit=False)
            contacto.organizacion = organizacion
            contacto.save()

            if request.FILES.getlist("archivos"):

                for archivo in request.FILES.getlist("archivos"):

                    ArchivoContacto.objects.create(
                        contacto=contacto,
                        archivo=archivo
                    )

            return redirect("contactos_list", slug=slug)

    else:

        form = ContactoForm()

    return render(
        request,
        "core/contacto_form.html",
        {
            "form": form,
            "slug": slug
        }
    )

@tenant_login_required
def contacto_update(request, slug, pk):

    contacto = get_object_or_404(
        Contacto,
        pk=pk,
        organizacion=request.organizacion
    )

    if request.method == "POST":

        form = ContactoForm(
            request.POST,
            request.FILES,
            instance=contacto,
            organizacion=request.organizacion
        )

        if form.is_valid():

            contacto = form.save()

            # subir nuevos archivos
            if request.FILES.getlist("archivos"):

                for archivo in request.FILES.getlist("archivos"):

                    ArchivoContacto.objects.create(
                        contacto=contacto,
                        archivo=archivo
                    )

            return redirect("contacto_detail", slug=slug, pk=contacto.id)

    else:

        form = ContactoForm(
            instance=contacto,
            organizacion=request.organizacion
        )

    return render(
        request,
        "core/contacto_form.html",
        {
            "form": form,
            "contacto": contacto,
            "slug": slug
        }
    )

@tenant_login_required
def archivo_contacto_delete(request, slug, pk):

    archivo = get_object_or_404(
        ArchivoContacto,
        pk=pk,
        contacto__organizacion=request.organizacion
    )

    contacto_id = archivo.contacto.id

    archivo.delete()

    return redirect("contacto_update", slug=slug, pk=contacto_id)

@tenant_login_required
@require_POST
def tipo_contacto_ajax_create(request, slug):

    nombre = request.POST.get("nombre")

    tipo = TipoContacto.objects.create(
        nombre=nombre,
        organizacion=request.organizacion
    )

    return JsonResponse({
        "id": tipo.id,
        "nombre": tipo.nombre
    })

@tenant_login_required
def contacto_detail(request, slug, pk):

    contacto = get_object_or_404(
        Contacto,
        pk=pk,
        organizacion=request.organizacion
    )

    return render(
        request,
        "core/contacto_detail.html",
        {
            "contacto": contacto,
            "slug": slug
        }
    )

@tenant_login_required
def inventario_list(request, slug):

    organizacion = request.organizacion

    items = ItemInventario.objects.filter(
        organizacion=organizacion
    ).select_related("categoria")

    return render(
        request,
        "core/inventario_list.html",
        {
            "items": items,
            "slug": slug
        }
    )


@tenant_login_required
def inventario_list(request, slug):

    organizacion = request.organizacion

    q = request.GET.get("q")
    categoria = request.GET.get("categoria")

    items = ItemInventario.objects.filter(
        organizacion=organizacion
    ).select_related("categoria")

    if q:
        items = items.filter(
            nombre__icontains=q
        ) | items.filter(
            descripcion__icontains=q
        ) | items.filter(
            ubicacion__icontains=q
        )

    if categoria:
        items = items.filter(categoria_id=categoria)

    categorias = CategoriaInventario.objects.filter(
        organizacion=organizacion
    )

    return render(
        request,
        "core/inventario_list.html",
        {
            "items": items,
            "categorias": categorias,
            "slug": slug
        }
    )

@tenant_login_required
@require_POST
def categoria_inventario_ajax_create(request, slug):

    nombre = request.POST.get("nombre")

    categoria = CategoriaInventario.objects.create(
        nombre=nombre,
        organizacion=request.organizacion
    )

    return JsonResponse({
        "id": categoria.id,
        "nombre": categoria.nombre
    })

@tenant_login_required
def inventario_create(request, slug):

    organizacion = request.organizacion

    if request.method == "POST":

        form = ItemInventarioForm(request.POST)

        if form.is_valid():

            item = form.save(commit=False)
            item.organizacion = organizacion
            item.save()

            if request.FILES.getlist("archivos"):

                for archivo in request.FILES.getlist("archivos"):

                    ArchivoInventario.objects.create(
                        item=item,
                        archivo=archivo,
                        es_imagen=archivo.content_type.startswith("image")
                    )

            return redirect("inventario_list", slug=slug)

    else:

        form = ItemInventarioForm()

    return render(
        request,
        "core/inventario_form.html",
        {
            "form": form,
            "slug": slug
        }
    )

@tenant_login_required
def inventario_update(request, slug, pk):

    item = get_object_or_404(
        ItemInventario,
        pk=pk,
        organizacion=request.organizacion
    )

    if request.method == "POST":

        form = ItemInventarioForm(
            request.POST,
            instance=item
        )

        if form.is_valid():

            form.save()

            if request.FILES.getlist("archivos"):

                for archivo in request.FILES.getlist("archivos"):

                    ArchivoInventario.objects.create(
                        item=item,
                        archivo=archivo,
                        es_imagen=archivo.content_type.startswith("image")
                    )

            return redirect("inventario_detail", slug=slug, pk=item.id)

    else:

        form = ItemInventarioForm(instance=item)

    return render(
        request,
        "core/inventario_form.html",
        {
            "form": form,
            "item": item,
            "slug": slug
        }
    )

@tenant_login_required
def inventario_detail(request, slug, pk):

    item = get_object_or_404(
        ItemInventario,
        pk=pk,
        organizacion=request.organizacion
    )

    return render(
        request,
        "core/inventario_detail.html",
        {
            "item": item,
            "slug": slug
        }
    )

@tenant_login_required
def inventario_delete(request, slug, pk):

    item = get_object_or_404(
        ItemInventario,
        pk=pk,
        organizacion=request.organizacion
    )

    item.delete()

    return redirect("inventario_list", slug=slug)

@tenant_login_required
def archivo_inventario_delete(request, slug, pk):

    archivo = get_object_or_404(
        ArchivoInventario,
        pk=pk,
        item__organizacion=request.organizacion
    )

    item_id = archivo.item.id

    archivo.delete()

    return redirect("inventario_detail", slug=slug, pk=item_id)

from core.exports import export_queryset_to_excel
from .models import Contacto


@tenant_login_required
def contactos_export(request):

    queryset = Contacto.objects.filter(
        organizacion=request.organizacion
    )

    fields = [
        "nombre",
        "email",
        "telefono",
        "empresa",
        "notas",
    ]

    headers = [
        "Nombre",
        "Email",
        "Teléfono",
        "Empresa",
        "Notas",
    ]

    return export_queryset_to_excel(
        queryset,
        fields,
        headers,
        "contactos",
    )

from core.exports import export_queryset_to_excel
from .models import Inventario


@tenant_login_required
def inventario_export(request):

    queryset = Inventario.objects.filter(
        organizacion=request.organizacion
    )

    fields = [
        "nombre",
        "categoria",
        "cantidad",
        "ubicacion",
        "estado",
        "descripcion",
    ]

    headers = [
        "Nombre",
        "Categoría",
        "Cantidad",
        "Ubicación",
        "Estado",
        "Descripción",
    ]

    return export_queryset_to_excel(
        queryset,
        fields,
        headers,
        "inventario",
    )

from django.shortcuts import redirect
from django.contrib import messages
import openpyxl


@tenant_login_required
def contactos_import(request):

    if request.method == "POST":

        file = request.FILES["archivo"]

        rows = read_excel(file)

        for row in rows:

            Contacto.objects.create(
                organizacion=request.organizacion,
                nombre=row[0],
                email=row[1],
                telefono=row[2],
                empresa=row[3],
                notas=row[4],
            )

        messages.success(request, "Contactos importados correctamente")

        return redirect("contactos_list")

    return render(
        request,
        "contactos/contactos_import.html",
    )

from core.excel_templates import generate_template


@tenant_login_required
def contactos_template(request):

    headers = [
        "Nombre",
        "Email",
        "Telefono",
        "Empresa",
        "Notas",
    ]

    return generate_template(headers, "plantilla_contactos")

from core.decorators import tenant_login_required
from core.excel import export_model_to_excel


@tenant_login_required
def excel_export(request, slug, model):

    return export_model_to_excel(
        request,
        "asociaciones",   # app_label
        model             # model_name
    )

import openpyxl
from django.http import HttpResponse
from asociaciones.models import Socio
from core.decorators import tenant_login_required


@tenant_login_required
def excel_socios_deuda(request, slug):

    socios = Socio.objects.filter(
        organizacion=request.organizacion,
        activo=True
    )

    socios_con_deuda = [
        socio for socio in socios
        if socio.total_deuda() > 0
    ]

    total_deuda = sum(s.total_deuda() for s in socios_con_deuda)

    wb = openpyxl.Workbook()

    # ─────────────────────────────
    # HOJA 1 — SOCIOS CON DEUDA
    # ─────────────────────────────

    ws = wb.active
    ws.title = "Socios con deuda"

    ws.append([
        "Nombre",
        "Familia",
        "Teléfono",
        "Email",
        "Deuda (€)"
    ])

    for socio in socios_con_deuda:

        familia = socio.familia.nombre if socio.familia else ""

        ws.append([
            f"{socio.nombre} {socio.apellidos}",
            familia,
            socio.telefono or "",
            socio.email or "",
            socio.total_deuda()
        ])

    # ─────────────────────────────
    # HOJA 2 — RESUMEN TESORERÍA
    # ─────────────────────────────

    ws_resumen = wb.create_sheet("Resumen")

    ws_resumen.append(["Indicador", "Valor"])
    ws_resumen.append(["Socios con deuda", len(socios_con_deuda)])
    ws_resumen.append(["Deuda total", f"{total_deuda} €"])

    # ─────────────────────────────
    # HOJA 3 — MENSAJE RECORDATORIO
    # ─────────────────────────────

    ws_msg = wb.create_sheet("Mensaje recordatorio")

    mensaje = [
        "Hola 👋",
        "",
        "Te recordamos que tienes pendiente la cuota anual de la asociación.",
        "",
        "Puedes realizar el pago en los próximos días.",
        "",
        "Si ya has realizado el pago, ignora este mensaje.",
        "",
        "Muchas gracias 🙂"
    ]

    for linea in mensaje:
        ws_msg.append([linea])

    # ─────────────────────────────
    # HOJA 4 — WHATSAPP
    # ─────────────────────────────

    ws_wp = wb.create_sheet("WhatsApp")

    ws_wp.append([
        "Nombre",
        "Teléfono",
        "Abrir WhatsApp"
    ])

    for socio in socios_con_deuda:

        telefono = socio.telefono or ""

        telefono_limpio = telefono.replace(" ", "").replace("+", "")

        if telefono_limpio and not telefono_limpio.startswith("34"):
            telefono_limpio = "34" + telefono_limpio

        link = f"https://wa.me/{telefono_limpio}" if telefono_limpio else ""

        ws_wp.append([
            f"{socio.nombre} {socio.apellidos}",
            telefono,
            link
        ])
        
    # ─────────────────────────────
    # DESCARGA
    # ─────────────────────────────

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="socios_deuda.xlsx"'
    formatear_excel(ws)
    wb.save(response)

    return response

def ajustar_columnas(ws):

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.db.models import Sum
from asociaciones.models import Pago, Gasto, Actividad, Patrocinador
from core.decorators import tenant_login_required


# ─────────────────────────────
# FUNCIONES AUXILIARES
# ─────────────────────────────

def poner_cabecera(ws):
    """Poner cabecera en negrita y congelar fila"""
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"


def ajustar_columnas(ws):
    """Ajustar automáticamente el ancho de columnas"""

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2


# ─────────────────────────────
# INFORME ANUAL
# ─────────────────────────────

@tenant_login_required
def informe_anual(request, slug):

    año = request.GET.get("anio")

    if not año:
        from datetime import date
        año = date.today().year
    else:
        año = int(año)

    # ─────────────────────────────
    # DATOS FINANCIEROS
    # ─────────────────────────────

    pagos = Pago.objects.filter(
        organizacion=request.organizacion,
        fecha__year=año
    )

    ingresos_cuotas = pagos.filter(
        cuota__isnull=False
    ).aggregate(total=Sum("importe"))["total"] or 0

    ingresos_actividades = pagos.filter(
        actividad__isnull=False
    ).aggregate(total=Sum("importe"))["total"] or 0

    ingresos_patrocinios = Patrocinador.objects.filter(
        año=año
    ).aggregate(total=Sum("aportacion"))["total"] or 0

    gastos = Gasto.objects.filter(
        fecha__year=año
    )

    total_gastos = gastos.aggregate(total=Sum("importe"))["total"] or 0

    total_ingresos = ingresos_cuotas + ingresos_actividades + ingresos_patrocinios
    beneficio = total_ingresos - total_gastos

    wb = openpyxl.Workbook()

    # ─────────────────────────────
    # HOJA 1 — RESUMEN
    # ─────────────────────────────

    ws = wb.active
    ws.title = "Resumen anual"

    ws.append(["Concepto", "Importe"])

    ws.append(["Ingresos cuotas", ingresos_cuotas])
    ws.append(["Ingresos actividades", ingresos_actividades])
    ws.append(["Ingresos patrocinadores", ingresos_patrocinios])
    ws.append(["Ingresos totales", total_ingresos])
    ws.append(["Gastos totales", total_gastos])
    ws.append(["Resultado final", beneficio])

    poner_cabecera(ws)

    # ─────────────────────────────
    # HOJA 2 — INGRESOS
    # ─────────────────────────────

    ws_ingresos = wb.create_sheet("Ingresos")

    ws_ingresos.append([
        "Fecha",
        "Socio",
        "Familia",
        "Concepto",
        "Actividad",
        "Importe",
        "Método"
    ])

    for pago in pagos.select_related("socio", "actividad", "familia", "cuota"):

        socio = str(pago.socio) if pago.socio else ""
        familia = pago.familia.nombre if pago.familia else ""
        actividad = pago.actividad.nombre if pago.actividad else ""

        # Determinar concepto del pago
        if pago.cuota:
            concepto = f"Cuota {pago.cuota.año}"
        elif pago.actividad:
            concepto = "Actividad"
        elif pago.familia:
            concepto = "Pago familiar"
        else:
            concepto = "Otro ingreso"

        ws_ingresos.append([
            pago.fecha,
            socio,
            familia,
            concepto,
            actividad,
            pago.importe,
            pago.metodo
        ])

    poner_cabecera(ws_ingresos)

    # ─────────────────────────────
    # HOJA 3 — GASTOS
    # ─────────────────────────────

    ws_gastos = wb.create_sheet("Gastos")

    ws_gastos.append([
        "Fecha",
        "Concepto",
        "Actividad",
        "Importe"
    ])

    for gasto in gastos.select_related("actividad"):

        actividad = gasto.actividad.nombre if gasto.actividad else ""

        ws_gastos.append([
            gasto.fecha,
            gasto.concepto,
            actividad,
            gasto.importe
        ])

    poner_cabecera(ws_gastos)

    # ─────────────────────────────
    # HOJA 4 — ACTIVIDADES
    # ─────────────────────────────

    ws_act = wb.create_sheet("Actividades")

    ws_act.append([
        "Actividad",
        "Ingresos",
        "Gastos",
        "Beneficio"
    ])

    actividades = Actividad.objects.filter(
        organizacion=request.organizacion
    ).order_by("fecha")

    for actividad in actividades:

        ingresos = actividad.ingresos_totales()
        gastos_act = actividad.total_gastos()
        beneficio_act = actividad.beneficio()

        ws_act.append([
            actividad.nombre,
            ingresos,
            gastos_act,
            beneficio_act
        ])

    poner_cabecera(ws_act)

    # ─────────────────────────────
    # HOJA 5 — INGRESOS POR ACTIVIDAD
    # ─────────────────────────────

    ws_detalle = wb.create_sheet("Ingresos por actividad")

    ws_detalle.append([
        "Actividad",
        "Fecha",
        "Participantes",
        "Ingresos",
        "Gastos",
        "Beneficio"
    ])

    for actividad in actividades:

        participantes = actividad.inscripciones.count()

        ingresos = actividad.ingresos_totales()
        gastos_act = actividad.total_gastos()
        beneficio_act = actividad.beneficio()

        ws_detalle.append([
            actividad.nombre,
            actividad.fecha,
            participantes,
            ingresos,
            gastos_act,
            beneficio_act
        ])

    poner_cabecera(ws_detalle)

    # ─────────────────────────────
    # AJUSTAR COLUMNAS
    # ─────────────────────────────

    formatear_excel(ws)
    formatear_excel(ws_ingresos)
    formatear_excel(ws_gastos)
    formatear_excel(ws_act)
    formatear_excel(ws_detalle)

    # ─────────────────────────────
    # DESCARGA
    # ─────────────────────────────

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"informe_asociacion_{año}.xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    formatear_excel(ws)
    wb.save(response)

    return response

from django.shortcuts import redirect, render
from django.contrib import messages
from django.db import transaction

from asociaciones.models import Socio, Familia
from asociaciones.imports import read_excel

@tenant_login_required
def importar_socios_excel(request, slug):

    if request.method == "POST":

        archivo = request.FILES.get("archivo")

        if not archivo:
            messages.error(request, "Debes subir un archivo Excel")
            return redirect("socios_list", slug=slug)

        try:
            filas = read_excel(archivo)
        except Exception:
            messages.error(request, "El archivo no es un Excel válido")
            return redirect("socios_list", slug=slug)

        errores = []

        # -----------------------------
        # VALIDAR TODO EL ARCHIVO
        # -----------------------------

        for fila in filas:

            if not fila["nombre"]:
                errores.append(
                    f"Fila {fila['fila']}: El nombre es obligatorio"
                )

        # CANCELAR SI HAY ERRORES

        if errores:

            for e in errores[:10]:
                messages.error(request, e)

            messages.error(
                request,
                f"Importación cancelada. {len(errores)} errores encontrados."
            )

            return redirect("socios_list", slug=slug)

        # -----------------------------
        # IMPORTAR (TRANSACCIÓN)
        # -----------------------------

        creados = 0

        with transaction.atomic():

            for fila in filas:

                familia = None

                if fila["familia_nombre"]:

                    familia, _ = Familia.objects.get_or_create(
                        organizacion=request.organizacion,
                        nombre=fila["familia_nombre"]
                    )

                Socio.objects.create(
                    organizacion=request.organizacion,
                    nombre=fila["nombre"],
                    apellidos=fila["apellidos"],
                    fecha_nacimiento=fila["fecha_nacimiento"],
                    familia=familia,
                    telefono=fila["telefono"],
                    email=fila["email"]
                )

                creados += 1

        messages.success(request, f"{creados} socios importados correctamente")

    return redirect("socios_list", slug=slug)

@tenant_login_required
def cuota_update(request, slug, pk):

    cuota = get_object_or_404(
        Cuota,
        pk=pk,
        organizacion=request.organizacion
    )

    form = CuotaForm(request.POST or None, instance=cuota)
    
    if cuota.estado != "borrador":
        for field in form.fields:
            form.fields[field].disabled = True

    if form.is_valid():
        form.save()
        return redirect("cuotas_list", slug=slug)

    return render(
        request,
        "core/cuota_form.html",
        {
            "form": form,
            "slug": slug,
            "cuota": cuota
        }
    )

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@tenant_login_required
def cuota_delete(request, slug, pk):

    cuota = get_object_or_404(
        Cuota,
        pk=pk,
        organizacion=request.organizacion
    )

    # Seguridad: solo borrar borradores
    if cuota.estado != "borrador":
        messages.error(request, "Solo se pueden eliminar cuotas en borrador.")
        return redirect("cuotas_list", slug=slug)

    cuota.delete()

    messages.success(request, "Cuota eliminada correctamente.")

    return redirect("cuotas_list", slug=slug)

from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.lib import colors

from asociaciones.models import Actividad, Inscripcion


def actividad_lista_asistencia(request, slug, pk):

    actividad = get_object_or_404(
        Actividad,
        pk=pk,
        organizacion=request.organizacion
    )

    inscripciones = (
        Inscripcion.objects
        .filter(actividad=actividad)
        .select_related("socio")
        .order_by("socio__apellidos")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="lista_{actividad.nombre}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=20,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph(f"<b>{actividad.nombre}</b>", styles["Title"]))
    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"Fecha: {actividad.fecha} &nbsp;&nbsp;&nbsp; "
            f"Plazas: {inscripciones.count()} / {actividad.cupo_maximo}",
            styles["Normal"]
        )
    )

    elements.append(Spacer(1, 20))

    # Responsable
    elements.append(Paragraph("Responsable actividad: ______________________________", styles["Normal"]))
    elements.append(Paragraph("Teléfono emergencia: ______________________________", styles["Normal"]))

    elements.append(Spacer(1, 20))

    # Tabla
    data = [["Nombre y apellidos", "Edad", "Teléfono", "Pagado", "Asistencia"]]

    for inscripcion in inscripciones:

        socio = inscripcion.socio

        nombre = f"{socio.nombre} {socio.apellidos}"
        edad = socio.edad() if socio.edad() else ""
        telefono = socio.telefono or ""
        pagado = "Sí" if inscripcion.pagado else "No"

        data.append([nombre, edad, telefono, pagado, ""])

    table = Table(data, colWidths=[7 * cm, 2 * cm, 4 * cm, 2 * cm, 3 * cm])

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ])
    )

    elements.append(table)

    elements.append(Spacer(1, 30))

    elements.append(Paragraph("Observaciones:", styles["Normal"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("__________________________________________________________________________", styles["Normal"]))
    elements.append(Paragraph("__________________________________________________________________________", styles["Normal"]))
    elements.append(Paragraph("__________________________________________________________________________", styles["Normal"]))

    doc.build(elements)

    return response

import openpyxl
from urllib.parse import quote
from django.http import HttpResponse
from asociaciones.models import Socio
from core.excel import formatear_excel


def export_socios_deuda(request, slug):

    socios = Socio.objects.filter(
        organizacion=request.organizacion,
        activo=True
    )

    wb = openpyxl.Workbook()

    # ==========================
    # HOJA 1 — SOCIOS DEUDA
    # ==========================

    ws = wb.active
    ws.title = "Socios con deuda"

    ws.append([
        "Nombre",
        "Teléfono",
        "Email",
        "Deuda (€)",
        "WhatsApp"
    ])

    total_deuda = 0
    total_deudores = 0

    for socio in socios:

        deuda = socio.total_deuda()

        if deuda <= 0:
            continue

        total_deudores += 1
        total_deuda += deuda

        nombre = f"{socio.nombre} {socio.apellidos}"
        telefono = socio.telefono or ""
        email = socio.email or ""

        mensaje = f"Hola {socio.nombre}, te recordamos que tienes una cuota pendiente de {deuda} € con la asociación. Gracias."

        mensaje_url = quote(mensaje)

        if telefono:
            whatsapp = f"https://wa.me/{telefono}?text={mensaje_url}"
        else:
            whatsapp = ""

        ws.append([
            nombre,
            telefono,
            email,
            deuda,
            whatsapp
        ])

    formatear_excel(ws)

    # ==========================
    # HOJA 2 — MENSAJES WHATSAPP
    # ==========================

    ws2 = wb.create_sheet("Mensajes WhatsApp")

    ws2.append([
        "Nombre",
        "Teléfono",
        "Mensaje"
    ])

    for socio in socios:

        deuda = socio.total_deuda()

        if deuda <= 0:
            continue

        nombre = f"{socio.nombre} {socio.apellidos}"
        telefono = socio.telefono or ""

        mensaje = (
            f"Hola {socio.nombre},\n\n"
            f"Te recordamos que tienes pendiente el pago de {deuda} €.\n\n"
            "Muchas gracias."
        )

        ws2.append([
            nombre,
            telefono,
            mensaje
        ])

    formatear_excel(ws2)

    # ==========================
    # HOJA 3 — RESUMEN
    # ==========================

    ws3 = wb.create_sheet("Resumen")

    ws3.append(["Concepto", "Valor"])
    ws3.append(["Socios con deuda", total_deudores])
    ws3.append(["Deuda total", total_deuda])

    formatear_excel(ws3)

    # ==========================
    # DESCARGA
    # ==========================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="socios_deuda.xlsx"'

    wb.save(response)

    return response
